#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Planejador de Escalas de Turismo - Interface Tkinter

Start:
    python tour_assignment_gui.py

Banco de dados:
    tour_assignment_data.json deve estar na mesma pasta.

Pacotes opcionais:
    pip install scipy numpy openpyxl

- scipy/numpy: otimização exata da escala diária
- openpyxl: exportação em .xlsx
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk


APP_TITLE = "Planejador de Escalas de Turismo"
APP_VERSION = "1.0"
DEFAULT_DATA_FILE = "tour_assignment_data.json"
DATE_RE = re.compile(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{4})$")
TIME_RE = re.compile(r"^\d{1,2}:\d{2}$")

RESULT_HEADERS = [
    "Escala",
    "Entrada/Saída",
    "O.S.",
    "Arquivo",
    "Pax",
    "Pax/Grupo",
    "Chegada",
    "Saída",
    "Voo de saída",
    "Hotel",
    "Idioma",
    "Guia",
    "Veículo / Motorista",
    "Agência",
    "Observação",
]

LANGUAGE_SUGGESTIONS = ["alemão", "espanhol", "francês", "inglês", "italiano", "português"]


# ---------------------------------------------------------------------------
# Allgemeine Hilfsfunktionen
# ---------------------------------------------------------------------------


def app_directory() -> Path:
    """Pasta do arquivo .py ou do executável iniciado."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def data_path() -> Path:
    return app_directory() / DEFAULT_DATA_FILE


def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def int_or_none(value: Any) -> Optional[int]:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(str(value).strip())
    except ValueError:
        return None


def bool_to_text(value: Any) -> str:
    return "Sim" if bool(value) else "Não"


def split_csv_text(value: str) -> List[str]:
    return [part.strip() for part in (value or "").split(",") if part.strip()]


def join_csv(values: Iterable[Any]) -> str:
    return ", ".join(str(v) for v in values if str(v).strip())


def normalize_escala(value: Any) -> str:
    """Normaliza DD-MM-YYYY, DD.MM.YYYY ou DD/MM/YYYY para DD-MM-YYYY."""
    text = str(value or "").strip()
    if not text:
        return ""
    match = DATE_RE.match(text)
    if not match:
        return text
    day, month, year = match.groups()
    return f"{int(day):02d}-{int(month):02d}-{year}"


def normalize_dot_date(value: Any) -> str:
    """Normaliza datas para DD.MM.YYYY, compatível com os campos antigos dos dicionários."""
    text = str(value or "").strip()
    if not text:
        return ""
    match = DATE_RE.match(text)
    if not match:
        return text
    day, month, year = match.groups()
    return f"{int(day):02d}.{int(month):02d}.{year}"


def escala_to_dot(value: Any) -> str:
    return normalize_dot_date(normalize_escala(value))


def is_valid_date_like(value: Any) -> bool:
    return bool(DATE_RE.match(str(value or "").strip()))


def date_sort_key(value: Any) -> Tuple[int, int, int, str]:
    text = normalize_escala(value)
    match = DATE_RE.match(text)
    if not match:
        return (9999, 99, 99, text)
    day, month, year = match.groups()
    return (int(year), int(month), int(day), text)


def compact_date(value: Any) -> str:
    text = normalize_escala(value)
    match = DATE_RE.match(text)
    if not match:
        return re.sub(r"\W+", "", text) or "semdata"
    day, month, year = match.groups()
    return f"{year}{int(month):02d}{int(day):02d}"


def time_sort_key(value: Any) -> Tuple[int, int, str]:
    text = str(value or "").strip()
    if not text:
        return (99, 99, "")
    parts = text.split(":")
    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
        return (int(parts[0]), int(parts[1]), text)
    return (99, 99, text)


def group_escala(group: Dict[str, Any]) -> str:
    return normalize_escala(
        group.get("escala")
        or group.get("voo_chegada_data")
        or group.get("voo_partida_data")
        or ""
    )


def ensure_group_defaults(group: Dict[str, Any]) -> Dict[str, Any]:
    group.setdefault("escala", group_escala(group))
    group["escala"] = normalize_escala(group.get("escala"))
    group.setdefault("in_out", True)
    group.setdefault("pax", 0)
    group.setdefault("pagou_pelo_guia_extra", False)
    group.setdefault("agencia", "")
    group.setdefault("O.S.", "")
    group.setdefault("numero_de_guias_externos", 0)
    group.setdefault("nomes_dos_passageiros", [])
    group.setdefault("nacionalidade", "")
    group.setdefault("hotel", "")
    group.setdefault("voo_chegada_existe", False)
    group.setdefault("voo_chegada_data", escala_to_dot(group.get("escala")))
    group.setdefault("voo_chegada_horario", "")
    group.setdefault("voo_chegada_numero", "")
    group.setdefault("voo_partida_existe", False)
    group.setdefault("voo_partida_data", "")
    group.setdefault("voo_partida_horario", "")
    group.setdefault("voo_partida_numero", "")
    group.setdefault("idioma", "")
    group.setdefault("file", "")
    group.setdefault("pedagios", "")
    return group


def new_empty_data() -> Dict[str, Any]:
    return {
        "version": 1,
        "settings": {"maximum_group_size_with_only_one_person": 7},
        "cars": {},
        "guides": {},
        "hotels": {},
        "tourist_groups": {},
    }


# ---------------------------------------------------------------------------
# Persistenz
# ---------------------------------------------------------------------------


class DataStore:
    def __init__(self, path: Path):
        self.path = path
        self.data: Dict[str, Any] = new_empty_data()
        self.load()

    def load(self) -> None:
        if self.path.exists():
            with self.path.open("r", encoding="utf-8") as f:
                loaded = json.load(f)
            if not isinstance(loaded, dict):
                raise ValueError("O banco de dados JSON precisa ser um objeto.")
            self.data = loaded
        else:
            self.data = new_empty_data()
            self.save()
        self._ensure_shape()

    def _ensure_shape(self) -> None:
        defaults = new_empty_data()
        for key, value in defaults.items():
            self.data.setdefault(key, value)
        self.data.setdefault("settings", {})
        self.data["settings"].setdefault("maximum_group_size_with_only_one_person", 7)
        for gid, group in list(self.groups.items()):
            if isinstance(group, dict):
                ensure_group_defaults(group)
            else:
                self.groups[gid] = ensure_group_defaults({})
        for name, guide in list(self.guides.items()):
            if not isinstance(guide, dict):
                self.guides[name] = {}
                guide = self.guides[name]
            guide.setdefault("languages", [])
            guide.setdefault("is_driver", False)
            guide.setdefault("is_guide", True)
            guide.setdefault("can_drive", [])
        for plate, car in list(self.cars.items()):
            if not isinstance(car, dict):
                self.cars[plate] = {}
                car = self.cars[plate]
            car.setdefault("frota", "")
            car.setdefault("marca_modelo", "")
            car.setdefault("renavam", "")
            car.setdefault("ano_modelo", "")
            car.setdefault("selo_foztrans", "")
            car.setdefault("numero_do_carro", None)
            car.setdefault("capacidade_min_de_pessoas", 1)
            car.setdefault("capacidade_max_de_pessoas", 3)
        for name, hotel in list(self.hotels.items()):
            if not isinstance(hotel, dict):
                self.hotels[name] = {}
                hotel = self.hotels[name]
            hotel.setdefault("adresse", "")
            hotel.setdefault("land", "")

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self.path.open("w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def backup_to(self, target: Path) -> None:
        self.save()
        shutil.copy2(self.path, target)

    @property
    def cars(self) -> Dict[str, Dict[str, Any]]:
        return self.data.setdefault("cars", {})

    @property
    def guides(self) -> Dict[str, Dict[str, Any]]:
        return self.data.setdefault("guides", {})

    @property
    def hotels(self) -> Dict[str, Dict[str, Any]]:
        return self.data.setdefault("hotels", {})

    @property
    def groups(self) -> Dict[str, Dict[str, Any]]:
        return self.data.setdefault("tourist_groups", {})

    @property
    def settings(self) -> Dict[str, Any]:
        return self.data.setdefault("settings", {})

    def all_escalas(self) -> List[str]:
        dates = {group_escala(group) for group in self.groups.values() if group_escala(group)}
        return sorted(dates, key=date_sort_key)

    def all_languages(self) -> List[str]:
        langs = set(LANGUAGE_SUGGESTIONS)
        for guide in self.guides.values():
            langs.update(str(v).strip() for v in guide.get("languages", []) if str(v).strip())
        for group in self.groups.values():
            if group.get("idioma"):
                langs.add(str(group["idioma"]).strip())
        return sorted(langs)


# ---------------------------------------------------------------------------
# Optimierung / Disposition
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Candidate:
    group_id: str
    plate: Optional[str]
    guide: str
    driver: str
    rented: bool
    people: frozenset
    cars: frozenset
    vehicle_score: int
    split_score: int
    person_score: int
    note: str = ""


class OptimizerError(RuntimeError):
    pass


def total_people(group: Dict[str, Any]) -> int:
    return safe_int(group.get("pax")) + safe_int(group.get("numero_de_guias_externos"))


def needs_two_people(group: Dict[str, Any], settings: Dict[str, Any]) -> bool:
    limit = safe_int(settings.get("maximum_group_size_with_only_one_person"), 7)
    return total_people(group) > limit or bool(group.get("pagou_pelo_guia_extra"))


def can_guide(guides: Dict[str, Dict[str, Any]], person_name: str, idioma: str) -> bool:
    person = guides.get(person_name, {})
    return bool(person.get("is_guide")) and idioma in person.get("languages", [])


def can_drive(guides: Dict[str, Dict[str, Any]], person_name: str, plate: str) -> bool:
    person = guides.get(person_name, {})
    return bool(person.get("is_driver")) and plate in person.get("can_drive", [])


def car_can_take(cars: Dict[str, Dict[str, Any]], plate: str, people_count: int) -> bool:
    car = cars.get(plate, {})
    return safe_int(car.get("capacidade_max_de_pessoas")) >= people_count


def vehicle_score(cars: Dict[str, Dict[str, Any]], plate: str, people_count: int) -> int:
    car = cars.get(plate, {})
    min_cap = safe_int(car.get("capacidade_min_de_pessoas"))
    max_cap = safe_int(car.get("capacidade_max_de_pessoas"))
    number = car.get("numero_do_carro")
    number_score = number if isinstance(number, int) else 99
    outside_min_range_penalty = 0 if min_cap <= people_count <= max_cap else 10_000
    return outside_min_range_penalty + (max_cap * 100) + (min_cap * 10) + number_score


def person_tiebreak_score(guides: Dict[str, Dict[str, Any]], guide: str, driver: str) -> int:
    score = 0
    if guide in guides:
        g = guides[guide]
        score += len(g.get("languages", [])) * 5
        score += len(g.get("can_drive", [])) // 10
    if driver in guides and driver != guide:
        d = guides[driver]
        score += 0 if (d.get("is_driver") and not d.get("is_guide")) else 100
        score += len(d.get("languages", [])) * 5
        score += len(d.get("can_drive", [])) // 10
    return score


def build_candidates_for_group(data: Dict[str, Any], group_id: str) -> List[Candidate]:
    cars = data.get("cars", {})
    guides = data.get("guides", {})
    settings = data.get("settings", {})
    groups = data.get("tourist_groups", {})
    group = groups[group_id]
    people_count = total_people(group)
    idioma = group.get("idioma", "")
    force_split = needs_two_people(group, settings)

    guide_names = sorted(
        [name for name in guides if can_guide(guides, name, idioma)],
        key=lambda name: (
            len(guides[name].get("languages", [])),
            len(guides[name].get("can_drive", [])),
            name,
        ),
    )

    candidates: List[Candidate] = []
    fitting_plates = sorted(
        [plate for plate in cars if car_can_take(cars, plate, people_count)],
        key=lambda plate: vehicle_score(cars, plate, people_count),
    )

    for plate in fitting_plates:
        car = cars[plate]
        min_cap = safe_int(car.get("capacidade_min_de_pessoas"))
        bigger_note = "VEÍCULO MAIOR QUE O MÍNIMO" if people_count < min_cap else ""
        drivers = sorted(
            [name for name in guides if can_drive(guides, name, plate)],
            key=lambda name: (
                not (guides[name].get("is_driver") and not guides[name].get("is_guide")),
                len(guides[name].get("can_drive", [])),
                len(guides[name].get("languages", [])),
                name,
            ),
        )

        if force_split:
            for guide in guide_names:
                for driver in drivers:
                    if guide == driver:
                        continue
                    note_parts = ["GUIA/MOTORISTA SEPARADOS"]
                    if bigger_note:
                        note_parts.append(bigger_note)
                    candidates.append(
                        Candidate(
                            group_id=group_id,
                            plate=plate,
                            guide=guide,
                            driver=driver,
                            rented=False,
                            people=frozenset([guide, driver]),
                            cars=frozenset([plate]),
                            vehicle_score=vehicle_score(cars, plate, people_count),
                            split_score=0,
                            person_score=person_tiebreak_score(guides, guide, driver),
                            note="; ".join(note_parts),
                        )
                    )
        else:
            for guide in guide_names:
                if guide in drivers:
                    candidates.append(
                        Candidate(
                            group_id=group_id,
                            plate=plate,
                            guide=guide,
                            driver=guide,
                            rented=False,
                            people=frozenset([guide]),
                            cars=frozenset([plate]),
                            vehicle_score=vehicle_score(cars, plate, people_count),
                            split_score=0,
                            person_score=person_tiebreak_score(guides, guide, guide),
                            note=bigger_note,
                        )
                    )

            for guide in guide_names:
                if guide not in drivers:
                    for driver in drivers:
                        if guide == driver:
                            continue
                        note_parts = ["GUIA/MOTORISTA SEPARADOS"]
                        if bigger_note:
                            note_parts.append(bigger_note)
                        candidates.append(
                            Candidate(
                                group_id=group_id,
                                plate=plate,
                                guide=guide,
                                driver=driver,
                                rented=False,
                                people=frozenset([guide, driver]),
                                cars=frozenset([plate]),
                                vehicle_score=vehicle_score(cars, plate, people_count),
                                split_score=500,
                                person_score=person_tiebreak_score(guides, guide, driver),
                                note="; ".join(note_parts),
                            )
                        )

    for guide in guide_names:
        candidates.append(
            Candidate(
                group_id=group_id,
                plate=None,
                guide=guide,
                driver="MOTORISTA TERCEIRIZADO",
                rented=True,
                people=frozenset([guide]),
                cars=frozenset(),
                vehicle_score=20_000 + (people_count * 100),
                split_score=0 if force_split else 100,
                person_score=person_tiebreak_score(guides, guide, "MOTORISTA TERCEIRIZADO"),
                note="ALUGAR VEÍCULO COM MOTORISTA",
            )
        )

    if not guide_names:
        candidates.append(
            Candidate(
                group_id=group_id,
                plate=None,
                guide="GUIA A DEFINIR",
                driver="MOTORISTA TERCEIRIZADO",
                rented=True,
                people=frozenset(),
                cars=frozenset(),
                vehicle_score=999_999,
                split_score=9_999,
                person_score=9_999,
                note=f"SEM GUIA INTERNO PARA O IDIOMA: {idioma}; ALUGAR/CONTRATAR",
            )
        )

    return sorted(
        candidates,
        key=lambda c: (c.rented, c.vehicle_score, c.split_score, c.person_score, c.plate or "ZZZ", c.guide, c.driver),
    )


def _selected_group_ids(data: Dict[str, Any], target_escala: str) -> List[str]:
    target = normalize_escala(target_escala)
    result = []
    for group_id, group in data.get("tourist_groups", {}).items():
        ensure_group_defaults(group)
        if group_escala(group) == target:
            result.append(group_id)
    return result


def solve_day_greedy(data: Dict[str, Any], target_escala: str) -> Tuple[Dict[str, Candidate], str]:
    """Fallback caso scipy/numpy não estejam instalados. Não é ótimo, mas evita sobreposições."""
    group_ids = _selected_group_ids(data, target_escala)
    group_ids.sort(
        key=lambda gid: (
            -total_people(data["tourist_groups"][gid]),
            time_sort_key(data["tourist_groups"][gid].get("voo_chegada_horario")),
            gid,
        )
    )
    used_people = set()
    used_cars = set()
    assignments: Dict[str, Candidate] = {}

    for gid in group_ids:
        candidates = build_candidates_for_group(data, gid)
        chosen = None
        for candidate in candidates:
            if candidate.people & used_people:
                continue
            if candidate.cars & used_cars:
                continue
            chosen = candidate
            break
        if chosen is None and candidates:
            # Último recurso: veículo alugado + guia interno ainda livre, se possível.
            for candidate in candidates:
                if candidate.rented and not (candidate.people & used_people):
                    chosen = candidate
                    break
        if chosen is None:
            group = data["tourist_groups"][gid]
            chosen = Candidate(
                group_id=gid,
                plate=None,
                guide="GUIA A DEFINIR",
                driver="MOTORISTA TERCEIRIZADO",
                rented=True,
                people=frozenset(),
                cars=frozenset(),
                vehicle_score=999_999,
                split_score=9_999,
                person_score=9_999,
                note=f"SEM RECURSO DISPONÍVEL; IDIOMA: {group.get('idioma', '')}",
            )
        assignments[gid] = chosen
        used_people.update(chosen.people)
        used_cars.update(chosen.cars)

    return assignments, "Fallback sem SciPy"


def solve_day_milp(data: Dict[str, Any], target_escala: str, time_limit: int = 60) -> Tuple[Dict[str, Candidate], str]:
    try:
        import numpy as np
        from scipy.optimize import Bounds, LinearConstraint, milp
        from scipy.sparse import csr_matrix, lil_matrix
    except Exception as exc:  # pragma: no cover - abhängig vom Zielsystem
        raise ImportError("scipy/numpy não instalados") from exc

    group_ids = _selected_group_ids(data, target_escala)
    if not group_ids:
        return {}, "MILP"

    candidates_by_group = {group_id: build_candidates_for_group(data, group_id) for group_id in group_ids}
    candidates = [candidate for group_id in group_ids for candidate in candidates_by_group[group_id]]
    if not candidates:
        return {}, "MILP"

    person_names = list(data.get("guides", {}).keys())
    car_plates = list(data.get("cars", {}).keys())

    row_count = len(group_ids) + len(person_names) + len(car_plates)
    matrix = lil_matrix((row_count, len(candidates)), dtype=float)
    lower = np.zeros(row_count)
    upper = np.ones(row_count)

    row = 0
    for group_id in group_ids:
        lower[row] = 1
        upper[row] = 1
        for col, candidate in enumerate(candidates):
            if candidate.group_id == group_id:
                matrix[row, col] = 1
        row += 1

    for person in person_names:
        for col, candidate in enumerate(candidates):
            if person in candidate.people:
                matrix[row, col] = 1
        row += 1

    for plate in car_plates:
        for col, candidate in enumerate(candidates):
            if plate in candidate.cars:
                matrix[row, col] = 1
        row += 1

    base_constraint = LinearConstraint(csr_matrix(matrix), lower, upper)

    rented_obj = np.array([1 if c.rented else 0 for c in candidates], dtype=float)
    vehicle_obj = np.array([c.vehicle_score for c in candidates], dtype=float)
    split_obj = np.array([c.split_score for c in candidates], dtype=float)
    person_obj = np.array([c.person_score for c in candidates], dtype=float)

    def _solve(objective, constraints):
        result = milp(
            c=objective.astype(float),
            integrality=np.ones(len(candidates), dtype=int),
            bounds=Bounds(0, 1),
            constraints=constraints,
            options={"time_limit": time_limit},
        )
        if not result.success:
            raise OptimizerError(f"O MILP não encontrou uma solução: {result.message}")
        return result

    result_1 = _solve(rented_obj, [base_constraint])
    min_rented = round(float(result_1.fun))
    rented_constraint = LinearConstraint(csr_matrix(rented_obj.reshape(1, -1)), [min_rented], [min_rented])

    result_2 = _solve(vehicle_obj, [base_constraint, rented_constraint])
    min_vehicle_score = round(float(result_2.fun))
    vehicle_constraint = LinearConstraint(
        csr_matrix(vehicle_obj.reshape(1, -1)),
        [min_vehicle_score],
        [min_vehicle_score],
    )

    final_obj = split_obj * 10_000 + person_obj
    result_3 = _solve(final_obj, [base_constraint, rented_constraint, vehicle_constraint])
    selected = [candidates[i] for i, value in enumerate(result_3.x) if value > 0.5]
    return {candidate.group_id: candidate for candidate in selected}, "MILP"


def solve_day(data: Dict[str, Any], target_escala: str, time_limit: int = 60) -> Tuple[Dict[str, Candidate], str]:
    try:
        return solve_day_milp(data, target_escala, time_limit=time_limit)
    except ImportError:
        return solve_day_greedy(data, target_escala)


def format_pax(group: Dict[str, Any]) -> str:
    pax = safe_int(group.get("pax"))
    extra = safe_int(group.get("numero_de_guias_externos"))
    return str(pax) if extra == 0 else f"{pax}+{extra}"


def format_flight(group: Dict[str, Any], kind: str) -> str:
    if not group.get(f"voo_{kind}_existe"):
        return ""
    number = group.get(f"voo_{kind}_numero") or ""
    hour = group.get(f"voo_{kind}_horario") or ""
    return f"{number} - {hour}" if number and hour else ""


def vehicle_driver_label(data: Dict[str, Any], candidate: Candidate, group: Dict[str, Any]) -> str:
    people_count = total_people(group)
    if candidate.rented or candidate.plate is None:
        return f"ALUGAR VEÍCULO {people_count} PAX - {candidate.driver}"

    car = data.get("cars", {}).get(candidate.plate, {})
    model = car.get("marca_modelo", "")
    car_number = car.get("numero_do_carro")
    number_part = f" {car_number}" if car_number is not None else ""
    return f"{model}{number_part} - {candidate.plate} - {candidate.driver}"


def build_output_rows(data: Dict[str, Any], assignments: Dict[str, Candidate], include_header: bool = True) -> List[List[Any]]:
    rows: List[List[Any]] = [RESULT_HEADERS[:]] if include_header else []
    groups = data.get("tourist_groups", {})

    def sort_key(group_id: str):
        group = groups[group_id]
        return (time_sort_key(group.get("voo_chegada_horario")), group.get("hotel", ""), group_id)

    for group_id in sorted(assignments, key=sort_key):
        group = groups[group_id]
        ensure_group_defaults(group)
        candidate = assignments[group_id]
        rows.append(
            [
                group_escala(group),
                bool_to_text(group.get("in_out")),
                group.get("O.S.", ""),
                group.get("file", ""),
                format_pax(group),
                " + ".join(group.get("nomes_dos_passageiros", [])),
                format_flight(group, "chegada"),
                group.get("voo_partida_data", "") if group.get("voo_partida_existe") else "",
                format_flight(group, "partida"),
                group.get("hotel", ""),
                group.get("idioma", ""),
                candidate.guide,
                vehicle_driver_label(data, candidate, group),
                group.get("agencia", ""),
                candidate.note,
            ]
        )
    return rows


# ---------------------------------------------------------------------------
# Tkinter-Helfer
# ---------------------------------------------------------------------------


class ScrollableFrame(ttk.Frame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)
        self.inner.bind("<Configure>", lambda event: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.window_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.scrollbar.grid(row=0, column=1, sticky="ns")
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.inner.bind("<Enter>", self._bind_mousewheel)
        self.inner.bind("<Leave>", self._unbind_mousewheel)

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.window_id, width=event.width)

    def _bind_mousewheel(self, _event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, _event):
        self.canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


def setup_tree(tree: ttk.Treeview, widths: Dict[str, int]) -> None:
    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=widths.get(col, 120), minwidth=40, stretch=True)


def clear_tree(tree: ttk.Treeview) -> None:
    for item in tree.get_children():
        tree.delete(item)


def selected_iid(tree: ttk.Treeview) -> Optional[str]:
    selection = tree.selection()
    return selection[0] if selection else None


def add_labeled_entry(parent, row: int, label: str, variable: tk.Variable, width: int = 28) -> ttk.Entry:
    ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=3)
    entry = ttk.Entry(parent, textvariable=variable, width=width)
    entry.grid(row=row, column=1, sticky="ew", pady=3)
    parent.columnconfigure(1, weight=1)
    return entry


def add_labeled_combo(parent, row: int, label: str, variable: tk.Variable, values: Sequence[str], width: int = 28) -> ttk.Combobox:
    ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=3)
    combo = ttk.Combobox(parent, textvariable=variable, values=list(values), width=width)
    combo.grid(row=row, column=1, sticky="ew", pady=3)
    parent.columnconfigure(1, weight=1)
    return combo


def show_error(title: str, exc: Exception) -> None:
    messagebox.showerror(title, str(exc))


# ---------------------------------------------------------------------------
# GUI-App
# ---------------------------------------------------------------------------


class TourAssignmentApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} {APP_VERSION}")
        self.geometry("1320x820")
        self.minsize(1100, 680)
        self.store = DataStore(data_path())
        self.current_rows: List[List[Any]] = []
        self.current_escala = tk.StringVar(value=self._default_escala())
        self.frames: Dict[str, Any] = {}
        self._build_menu()
        self._build_ui()
        self.set_status(f"Banco de dados: {self.store.path}")

    def _default_escala(self) -> str:
        dates = self.store.all_escalas()
        return dates[0] if dates else ""

    def _build_menu(self) -> None:
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=False)
        file_menu.add_command(label="Salvar dados", command=self.save_data)
        file_menu.add_command(label="Exportar backup do banco de dados...", command=self.export_database_backup)
        file_menu.add_command(label="Importar banco de dados de JSON...", command=self.import_database)
        file_menu.add_separator()
        file_menu.add_command(label="Sair", command=self.destroy)
        menubar.add_cascade(label="Arquivo", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=False)
        help_menu.add_command(label="Sobre", command=self.show_about)
        menubar.add_cascade(label="Ajuda", menu=help_menu)
        self.config(menu=menubar)

    def _build_ui(self) -> None:
        container = ttk.Frame(self, padding=8)
        container.grid(row=0, column=0, sticky="nsew")
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.notebook = ttk.Notebook(container)
        self.notebook.grid(row=0, column=0, sticky="nsew")
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)

        self.frames["planner"] = PlannerFrame(self.notebook, self)
        self.frames["groups"] = GroupsFrame(self.notebook, self)
        self.frames["cars"] = CarsFrame(self.notebook, self)
        self.frames["guides"] = GuidesFrame(self.notebook, self)
        self.frames["hotels"] = HotelsFrame(self.notebook, self)
        self.frames["settings"] = SettingsFrame(self.notebook, self)

        self.notebook.add(self.frames["planner"], text="Planejamento / Resultado")
        self.notebook.add(self.frames["groups"], text="Grupos")
        self.notebook.add(self.frames["cars"], text="Veículos")
        self.notebook.add(self.frames["guides"], text="Guias / Motoristas")
        self.notebook.add(self.frames["hotels"], text="Hotéis")
        self.notebook.add(self.frames["settings"], text="Configurações")

        self.status_var = tk.StringVar(value="Pronto")
        status = ttk.Label(container, textvariable=self.status_var, anchor="w")
        status.grid(row=1, column=0, sticky="ew", pady=(8, 0))

    def save_data(self) -> None:
        try:
            self.store.save()
            self.set_status("Dados salvos.")
        except Exception as exc:
            show_error("Falha ao salvar", exc)

    def refresh_all(self) -> None:
        for frame in self.frames.values():
            if hasattr(frame, "refresh"):
                frame.refresh()

    def refresh_references(self) -> None:
        for key in ("planner", "groups", "guides"):
            frame = self.frames.get(key)
            if frame is not None and hasattr(frame, "refresh_references"):
                frame.refresh_references()

    def set_status(self, text: str) -> None:
        self.status_var.set(text)

    def export_database_backup(self) -> None:
        target = filedialog.asksaveasfilename(
            title="Salvar backup do banco de dados",
            defaultextension=".json",
            filetypes=[("JSON", "*.json"), ("Todos os arquivos", "*.*")],
            initialfile="tour_assignment_data_backup.json",
        )
        if not target:
            return
        try:
            self.store.backup_to(Path(target))
            self.set_status(f"Backup salvo: {target}")
        except Exception as exc:
            show_error("Falha no backup", exc)

    def import_database(self) -> None:
        source = filedialog.askopenfilename(
            title="Importar banco de dados JSON",
            filetypes=[("JSON", "*.json"), ("Todos os arquivos", "*.*")],
        )
        if not source:
            return
        if not messagebox.askyesno("Substituir banco de dados", "O banco de dados atual será substituído. Continuar?"):
            return
        try:
            with open(source, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if not isinstance(loaded, dict):
                raise ValueError("O arquivo não contém um banco de dados válido.")
            self.store.data = loaded
            self.store._ensure_shape()
            self.store.save()
            self.current_rows = []
            self.current_escala.set(self._default_escala())
            self.refresh_all()
            self.set_status(f"Banco de dados importado: {source}")
        except Exception as exc:
            show_error("Falha na importação", exc)

    def show_about(self) -> None:
        messagebox.showinfo(
            "Sobre",
            f"{APP_TITLE} {APP_VERSION}\n\n"
            "Interface Tkinter para veículos, guias, motoristas, hotéis, grupos diários e exportação.\n"
            "Os dados são salvos em tour_assignment_data.json.",
        )


class PlannerFrame(ttk.Frame):
    def __init__(self, parent, app: TourAssignmentApp):
        super().__init__(parent, padding=8)
        self.app = app
        self.method_var = tk.StringVar(value="")
        self._build()
        self.refresh()

    def _build(self) -> None:
        top = ttk.Frame(self)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        ttk.Label(top, text="Escala:").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.date_combo = ttk.Combobox(top, textvariable=self.app.current_escala, width=14)
        self.date_combo.grid(row=0, column=1, sticky="w", padx=(0, 8))
        ttk.Button(top, text="Calcular escala", command=self.calculate).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(top, text="Exportar Excel", command=lambda: self.export_result("xlsx")).grid(row=0, column=3, padx=(0, 4))
        ttk.Button(top, text="Exportar CSV", command=lambda: self.export_result("csv")).grid(row=0, column=4, padx=4)
        ttk.Button(top, text="Exportar JSON", command=lambda: self.export_result("json")).grid(row=0, column=5, padx=4)
        ttk.Label(top, textvariable=self.method_var).grid(row=0, column=6, sticky="w", padx=(16, 0))
        top.columnconfigure(7, weight=1)

        table_frame = ttk.Frame(self)
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(table_frame, columns=RESULT_HEADERS, show="headings")
        widths = {
            "Escala": 88,
            "Entrada/Saída": 70,
            "O.S.": 80,
            "Arquivo": 80,
            "Pax": 70,
            "Pax/Grupo": 220,
            "Chegada": 130,
            "Saída": 100,
            "Voo de saída": 130,
            "Hotel": 240,
            "Idioma": 90,
            "Guia": 130,
            "Veículo / Motorista": 260,
            "Agência": 120,
            "Observação": 240,
        }
        setup_tree(self.tree, widths)
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

    def refresh_references(self) -> None:
        self.date_combo.configure(values=self.app.store.all_escalas())

    def refresh(self) -> None:
        self.refresh_references()

    def calculate(self) -> None:
        escala = normalize_escala(self.app.current_escala.get())
        if not escala or not is_valid_date_like(escala):
            messagebox.showwarning("Escala ausente", "Informe uma escala no formato DD-MM-YYYY.")
            return
        self.app.current_escala.set(escala)
        try:
            assignments, method = solve_day(self.app.store.data, escala)
            rows = build_output_rows(self.app.store.data, assignments, include_header=True)
            self.app.current_rows = rows
            self._populate(rows)
            rented = sum(1 for candidate in assignments.values() if candidate.rented)
            info = f"{len(assignments)} grupos planejados – método: {method}"
            if rented:
                info += f" – {rented} veículo(s) alugado(s)/solução(ões) externa(s)"
            self.method_var.set(info)
            self.app.set_status(info)
        except Exception as exc:
            show_error("Falha no planejamento", exc)

    def _populate(self, rows: List[List[Any]]) -> None:
        clear_tree(self.tree)
        for index, row in enumerate(rows[1:], start=1):
            tag = "rented" if row and "ALUGAR" in str(row[-1]).upper() else "normal"
            self.tree.insert("", "end", iid=str(index), values=row, tags=(tag,))
        self.tree.tag_configure("rented", background="#fff2cc")

    def export_result(self, kind: str) -> None:
        if not self.app.current_rows or len(self.app.current_rows) <= 1:
            messagebox.showwarning("Sem resultados", "Calcule a escala primeiro.")
            return
        escala = normalize_escala(self.app.current_escala.get()) or "escala"
        if kind == "xlsx":
            ext = ".xlsx"
            filetypes = [("Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        elif kind == "csv":
            ext = ".csv"
            filetypes = [("CSV", "*.csv"), ("Todos os arquivos", "*.*")]
        else:
            ext = ".json"
            filetypes = [("JSON", "*.json"), ("Todos os arquivos", "*.*")]
        target = filedialog.asksaveasfilename(
            title=f"Exportar resultado como {kind.upper()}",
            defaultextension=ext,
            filetypes=filetypes,
            initialfile=f"escala_{compact_date(escala)}{ext}",
        )
        if not target:
            return
        try:
            if kind == "xlsx":
                export_xlsx(self.app.current_rows, Path(target))
            elif kind == "csv":
                export_csv(self.app.current_rows, Path(target))
            else:
                export_json(self.app.current_rows, Path(target))
            self.app.set_status(f"Exportação salva: {target}")
        except Exception as exc:
            show_error("Falha na exportação", exc)


class CarsFrame(ttk.Frame):
    columns = [
        "Placa",
        "Frota",
        "Modelo",
        "Número",
        "Mín.",
        "Máx.",
        "Renavam",
        "Ano",
        "Selo",
    ]

    def __init__(self, parent, app: TourAssignmentApp):
        super().__init__(parent, padding=8)
        self.app = app
        self.current_plate: Optional[str] = None
        self.vars = {
            "plate": tk.StringVar(),
            "frota": tk.StringVar(),
            "marca_modelo": tk.StringVar(),
            "renavam": tk.StringVar(),
            "ano_modelo": tk.StringVar(),
            "selo_foztrans": tk.StringVar(),
            "numero_do_carro": tk.StringVar(),
            "capacidade_min_de_pessoas": tk.StringVar(),
            "capacidade_max_de_pessoas": tk.StringVar(),
        }
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.grid(row=0, column=0, sticky="nsew")

        left = ttk.Frame(paned)
        right = ttk.Frame(paned, padding=(12, 0, 0, 0))
        paned.add(left, weight=3)
        paned.add(right, weight=1)
        left.columnconfigure(0, weight=1)
        left.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(left, columns=self.columns, show="headings", selectmode="browse")
        setup_tree(self.tree, {"Placa": 100, "Modelo": 180, "Frota": 110, "Número": 70, "Mín.": 55, "Máx.": 55})
        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(left, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        form = ttk.LabelFrame(right, text="Editar veículo", padding=10)
        form.grid(row=0, column=0, sticky="nsew")
        right.columnconfigure(0, weight=1)
        add_labeled_entry(form, 0, "Placa", self.vars["plate"])
        add_labeled_entry(form, 1, "Frota", self.vars["frota"])
        add_labeled_entry(form, 2, "Marca/Modelo", self.vars["marca_modelo"])
        add_labeled_entry(form, 3, "Renavam", self.vars["renavam"])
        add_labeled_entry(form, 4, "Ano/Modelo", self.vars["ano_modelo"])
        add_labeled_entry(form, 5, "Selo Foztrans", self.vars["selo_foztrans"])
        add_labeled_entry(form, 6, "Número do carro", self.vars["numero_do_carro"])
        add_labeled_entry(form, 7, "Capacidade mín.", self.vars["capacidade_min_de_pessoas"])
        add_labeled_entry(form, 8, "Capacidade máx.", self.vars["capacidade_max_de_pessoas"])

        buttons = ttk.Frame(form)
        buttons.grid(row=9, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        ttk.Button(buttons, text="Novo", command=self.new_item).grid(row=0, column=0, padx=(0, 4))
        ttk.Button(buttons, text="Salvar", command=self.save_item).grid(row=0, column=1, padx=4)
        ttk.Button(buttons, text="Excluir", command=self.delete_item).grid(row=0, column=2, padx=4)

    def refresh(self) -> None:
        clear_tree(self.tree)
        for plate, car in sorted(self.app.store.cars.items()):
            values = [
                plate,
                car.get("frota", ""),
                car.get("marca_modelo", ""),
                car.get("numero_do_carro", ""),
                car.get("capacidade_min_de_pessoas", ""),
                car.get("capacidade_max_de_pessoas", ""),
                car.get("renavam", ""),
                car.get("ano_modelo", ""),
                car.get("selo_foztrans", ""),
            ]
            self.tree.insert("", "end", iid=plate, values=values)

    def on_select(self, _event=None) -> None:
        iid = selected_iid(self.tree)
        if not iid:
            return
        self.current_plate = iid
        car = self.app.store.cars[iid]
        self.vars["plate"].set(iid)
        for key in self.vars:
            if key == "plate":
                continue
            self.vars[key].set("" if car.get(key) is None else str(car.get(key, "")))

    def new_item(self) -> None:
        self.current_plate = None
        for var in self.vars.values():
            var.set("")
        self.vars["capacidade_min_de_pessoas"].set("1")
        self.vars["capacidade_max_de_pessoas"].set("3")
        self.tree.selection_remove(self.tree.selection())

    def save_item(self) -> None:
        old_plate = self.current_plate
        plate = self.vars["plate"].get().strip().upper()
        if not plate:
            messagebox.showwarning("Placa ausente", "Informe uma placa.")
            return
        if old_plate and plate != old_plate and plate in self.app.store.cars:
            messagebox.showwarning("Placa já existe", "Esta placa já existe.")
            return
        min_cap = safe_int(self.vars["capacidade_min_de_pessoas"].get(), -1)
        max_cap = safe_int(self.vars["capacidade_max_de_pessoas"].get(), -1)
        if min_cap < 0 or max_cap < min_cap:
            messagebox.showwarning("Verificar capacidade", "Informe capacidades mínima e máxima válidas.")
            return
        car = {
            "frota": self.vars["frota"].get().strip(),
            "marca_modelo": self.vars["marca_modelo"].get().strip(),
            "renavam": self.vars["renavam"].get().strip() or None,
            "ano_modelo": self.vars["ano_modelo"].get().strip(),
            "selo_foztrans": self.vars["selo_foztrans"].get().strip() or None,
            "numero_do_carro": int_or_none(self.vars["numero_do_carro"].get()),
            "capacidade_min_de_pessoas": min_cap,
            "capacidade_max_de_pessoas": max_cap,
        }
        if old_plate and plate != old_plate:
            self.app.store.cars.pop(old_plate, None)
            for guide in self.app.store.guides.values():
                guide["can_drive"] = [plate if p == old_plate else p for p in guide.get("can_drive", [])]
        self.app.store.cars[plate] = car
        self.current_plate = plate
        self.app.store.save()
        self.refresh()
        self.app.refresh_references()
        self.tree.selection_set(plate)
        self.app.set_status(f"Veículo salvo: {plate}")

    def delete_item(self) -> None:
        plate = self.current_plate or selected_iid(self.tree)
        if not plate:
            return
        if not messagebox.askyesno("Excluir veículo", f"Deseja realmente excluir o veículo {plate}?"):
            return
        self.app.store.cars.pop(plate, None)
        for guide in self.app.store.guides.values():
            guide["can_drive"] = [p for p in guide.get("can_drive", []) if p != plate]
        self.current_plate = None
        self.app.store.save()
        self.refresh()
        self.app.refresh_references()
        self.new_item()
        self.app.set_status(f"Veículo excluído: {plate}")


class GuidesFrame(ttk.Frame):
    columns = ["Nome", "Idiomas", "Motorista", "Guia", "Veículos"]

    def __init__(self, parent, app: TourAssignmentApp):
        super().__init__(parent, padding=8)
        self.app = app
        self.current_name: Optional[str] = None
        self.name_var = tk.StringVar()
        self.languages_var = tk.StringVar()
        self.is_driver_var = tk.BooleanVar(value=True)
        self.is_guide_var = tk.BooleanVar(value=True)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.grid(row=0, column=0, sticky="nsew")
        left = ttk.Frame(paned)
        right = ttk.Frame(paned, padding=(12, 0, 0, 0))
        paned.add(left, weight=3)
        paned.add(right, weight=2)
        left.columnconfigure(0, weight=1)
        left.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(left, columns=self.columns, show="headings", selectmode="browse")
        setup_tree(self.tree, {"Nome": 160, "Idiomas": 220, "Motorista": 70, "Guia": 70, "Veículos": 90})
        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        form = ttk.LabelFrame(right, text="Editar guia/motorista", padding=10)
        form.grid(row=0, column=0, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)
        add_labeled_entry(form, 0, "Nome", self.name_var)
        add_labeled_entry(form, 1, "Idiomas (vírgula)", self.languages_var)
        ttk.Checkbutton(form, text="É motorista", variable=self.is_driver_var).grid(row=2, column=0, columnspan=2, sticky="w", pady=3)
        ttk.Checkbutton(form, text="É guia", variable=self.is_guide_var).grid(row=3, column=0, columnspan=2, sticky="w", pady=3)
        ttk.Label(form, text="Pode dirigir:").grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 2))
        list_frame = ttk.Frame(form)
        list_frame.grid(row=5, column=0, columnspan=2, sticky="nsew")
        form.rowconfigure(5, weight=1)
        form.columnconfigure(1, weight=1)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        self.car_listbox = tk.Listbox(list_frame, selectmode="extended", height=14, exportselection=False)
        car_vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.car_listbox.yview)
        self.car_listbox.configure(yscrollcommand=car_vsb.set)
        self.car_listbox.grid(row=0, column=0, sticky="nsew")
        car_vsb.grid(row=0, column=1, sticky="ns")

        list_buttons = ttk.Frame(form)
        list_buttons.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Button(list_buttons, text="Selecionar todos", command=self.select_all_cars).grid(row=0, column=0, padx=(0, 4))
        ttk.Button(list_buttons, text="Limpar seleção", command=self.clear_car_selection).grid(row=0, column=1, padx=4)

        buttons = ttk.Frame(form)
        buttons.grid(row=7, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        ttk.Button(buttons, text="Novo", command=self.new_item).grid(row=0, column=0, padx=(0, 4))
        ttk.Button(buttons, text="Salvar", command=self.save_item).grid(row=0, column=1, padx=4)
        ttk.Button(buttons, text="Excluir", command=self.delete_item).grid(row=0, column=2, padx=4)

    def refresh_references(self) -> None:
        selected = set(self.selected_cars())
        self.car_listbox.delete(0, "end")
        for plate in sorted(self.app.store.cars):
            car = self.app.store.cars[plate]
            label = f"{plate} - {car.get('marca_modelo', '')}"
            self.car_listbox.insert("end", label)
            if plate in selected:
                self.car_listbox.selection_set(self.car_listbox.size() - 1)

    def refresh(self) -> None:
        self.refresh_references()
        clear_tree(self.tree)
        for name, guide in sorted(self.app.store.guides.items()):
            values = [
                name,
                join_csv(guide.get("languages", [])),
                bool_to_text(guide.get("is_driver")),
                bool_to_text(guide.get("is_guide")),
                len(guide.get("can_drive", [])),
            ]
            self.tree.insert("", "end", iid=name, values=values)

    def selected_cars(self) -> List[str]:
        plates = sorted(self.app.store.cars)
        result = []
        for idx in self.car_listbox.curselection():
            if 0 <= idx < len(plates):
                result.append(plates[idx])
        return result

    def set_selected_cars(self, plates: Sequence[str]) -> None:
        plate_set = set(plates)
        self.car_listbox.selection_clear(0, "end")
        for idx, plate in enumerate(sorted(self.app.store.cars)):
            if plate in plate_set:
                self.car_listbox.selection_set(idx)

    def select_all_cars(self) -> None:
        self.car_listbox.selection_set(0, "end")

    def clear_car_selection(self) -> None:
        self.car_listbox.selection_clear(0, "end")

    def on_select(self, _event=None) -> None:
        iid = selected_iid(self.tree)
        if not iid:
            return
        self.current_name = iid
        guide = self.app.store.guides[iid]
        self.name_var.set(iid)
        self.languages_var.set(join_csv(guide.get("languages", [])))
        self.is_driver_var.set(bool(guide.get("is_driver")))
        self.is_guide_var.set(bool(guide.get("is_guide")))
        self.set_selected_cars(guide.get("can_drive", []))

    def new_item(self) -> None:
        self.current_name = None
        self.name_var.set("")
        self.languages_var.set("")
        self.is_driver_var.set(True)
        self.is_guide_var.set(True)
        self.clear_car_selection()
        self.tree.selection_remove(self.tree.selection())

    def save_item(self) -> None:
        old_name = self.current_name
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Nome ausente", "Informe um nome.")
            return
        if old_name and name != old_name and name in self.app.store.guides:
            messagebox.showwarning("Nome já existe", "Este nome já existe.")
            return
        guide = {
            "languages": split_csv_text(self.languages_var.get()),
            "is_driver": bool(self.is_driver_var.get()),
            "is_guide": bool(self.is_guide_var.get()),
            "can_drive": self.selected_cars(),
        }
        if old_name and name != old_name:
            self.app.store.guides.pop(old_name, None)
        self.app.store.guides[name] = guide
        self.current_name = name
        self.app.store.save()
        self.refresh()
        self.app.refresh_references()
        self.tree.selection_set(name)
        self.app.set_status(f"Guia/motorista salvo: {name}")

    def delete_item(self) -> None:
        name = self.current_name or selected_iid(self.tree)
        if not name:
            return
        if not messagebox.askyesno("Excluir guia/motorista", f"Deseja realmente excluir {name}?"):
            return
        self.app.store.guides.pop(name, None)
        self.current_name = None
        self.app.store.save()
        self.refresh()
        self.new_item()
        self.app.set_status(f"Guia/motorista excluído: {name}")


class HotelsFrame(ttk.Frame):
    columns = ["Hotel", "País", "Endereço"]

    def __init__(self, parent, app: TourAssignmentApp):
        super().__init__(parent, padding=8)
        self.app = app
        self.current_name: Optional[str] = None
        self.name_var = tk.StringVar()
        self.land_var = tk.StringVar()
        self.address_text: tk.Text
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.grid(row=0, column=0, sticky="nsew")
        left = ttk.Frame(paned)
        right = ttk.Frame(paned, padding=(12, 0, 0, 0))
        paned.add(left, weight=3)
        paned.add(right, weight=1)
        left.columnconfigure(0, weight=1)
        left.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(left, columns=self.columns, show="headings", selectmode="browse")
        setup_tree(self.tree, {"Hotel": 280, "País": 110, "Endereço": 420})
        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        form = ttk.LabelFrame(right, text="Editar hotel", padding=10)
        form.grid(row=0, column=0, sticky="nsew")
        right.columnconfigure(0, weight=1)
        add_labeled_entry(form, 0, "Nome do hotel", self.name_var)
        add_labeled_entry(form, 1, "País", self.land_var)
        ttk.Label(form, text="Endereço").grid(row=2, column=0, sticky="nw", padx=(0, 8), pady=3)
        self.address_text = tk.Text(form, height=5, wrap="word")
        self.address_text.grid(row=2, column=1, sticky="nsew", pady=3)
        form.columnconfigure(1, weight=1)

        buttons = ttk.Frame(form)
        buttons.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        ttk.Button(buttons, text="Novo", command=self.new_item).grid(row=0, column=0, padx=(0, 4))
        ttk.Button(buttons, text="Salvar", command=self.save_item).grid(row=0, column=1, padx=4)
        ttk.Button(buttons, text="Excluir", command=self.delete_item).grid(row=0, column=2, padx=4)

    def refresh(self) -> None:
        clear_tree(self.tree)
        for name, hotel in sorted(self.app.store.hotels.items()):
            self.tree.insert("", "end", iid=name, values=[name, hotel.get("land", ""), hotel.get("adresse", "")])

    def on_select(self, _event=None) -> None:
        iid = selected_iid(self.tree)
        if not iid:
            return
        self.current_name = iid
        hotel = self.app.store.hotels[iid]
        self.name_var.set(iid)
        self.land_var.set(hotel.get("land", ""))
        self.address_text.delete("1.0", "end")
        self.address_text.insert("1.0", hotel.get("adresse", ""))

    def new_item(self) -> None:
        self.current_name = None
        self.name_var.set("")
        self.land_var.set("")
        self.address_text.delete("1.0", "end")
        self.tree.selection_remove(self.tree.selection())

    def save_item(self) -> None:
        old_name = self.current_name
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Hotel ausente", "Informe o nome do hotel.")
            return
        if old_name and name != old_name and name in self.app.store.hotels:
            messagebox.showwarning("Hotel já existe", "Este nome de hotel já existe.")
            return
        hotel = {
            "adresse": self.address_text.get("1.0", "end").strip(),
            "land": self.land_var.get().strip(),
        }
        if old_name and name != old_name:
            self.app.store.hotels.pop(old_name, None)
            for group in self.app.store.groups.values():
                if group.get("hotel") == old_name:
                    group["hotel"] = name
        self.app.store.hotels[name] = hotel
        self.current_name = name
        self.app.store.save()
        self.refresh()
        self.app.refresh_references()
        self.tree.selection_set(name)
        self.app.set_status(f"Hotel salvo: {name}")

    def delete_item(self) -> None:
        name = self.current_name or selected_iid(self.tree)
        if not name:
            return
        used = any(group.get("hotel") == name for group in self.app.store.groups.values())
        text = f"Deseja realmente excluir o hotel {name}?"
        if used:
            text += "\n\nObservação: este hotel é usado em grupos. As entradas dos grupos permanecerão como texto."
        if not messagebox.askyesno("Excluir hotel", text):
            return
        self.app.store.hotels.pop(name, None)
        self.current_name = None
        self.app.store.save()
        self.refresh()
        self.app.refresh_references()
        self.new_item()
        self.app.set_status(f"Hotel excluído: {name}")


class GroupsFrame(ttk.Frame):
    columns = ["ID", "Escala", "Entrada/Saída", "O.S.", "Arquivo", "Pax", "Extra", "Passageiros", "Hotel", "Idioma", "Chegada", "Partida", "Agência"]

    def __init__(self, parent, app: TourAssignmentApp):
        super().__init__(parent, padding=8)
        self.app = app
        self.current_group_id: Optional[str] = None
        self.vars = {
            "group_id": tk.StringVar(),
            "escala": tk.StringVar(),
            "in_out": tk.BooleanVar(value=True),
            "O.S.": tk.StringVar(),
            "file": tk.StringVar(),
            "agencia": tk.StringVar(),
            "pax": tk.StringVar(),
            "numero_de_guias_externos": tk.StringVar(),
            "pagou_pelo_guia_extra": tk.BooleanVar(value=False),
            "nacionalidade": tk.StringVar(),
            "hotel": tk.StringVar(),
            "idioma": tk.StringVar(),
            "pedagios": tk.StringVar(),
            "voo_chegada_existe": tk.BooleanVar(value=True),
            "voo_chegada_data": tk.StringVar(),
            "voo_chegada_numero": tk.StringVar(),
            "voo_chegada_horario": tk.StringVar(),
            "voo_partida_existe": tk.BooleanVar(value=False),
            "voo_partida_data": tk.StringVar(),
            "voo_partida_numero": tk.StringVar(),
            "voo_partida_horario": tk.StringVar(),
        }
        self.filter_var = tk.StringVar(value="Todas")
        self.passengers_text: tk.Text
        self._build()
        self.refresh()
        self.new_item()

    def _build(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        top = ttk.Frame(self)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Label(top, text="Filtrar Escala:").grid(row=0, column=0, padx=(0, 6))
        self.filter_combo = ttk.Combobox(top, textvariable=self.filter_var, width=16)
        self.filter_combo.grid(row=0, column=1, padx=(0, 8))
        self.filter_combo.bind("<<ComboboxSelected>>", lambda _e: self.refresh_tree())
        ttk.Button(top, text="Aplicar", command=self.refresh_tree).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(top, text="Novo grupo", command=self.new_item).grid(row=0, column=3, padx=(0, 8))
        ttk.Button(top, text="Copiar grupo", command=self.copy_item).grid(row=0, column=4)
        top.columnconfigure(5, weight=1)

        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.grid(row=1, column=0, sticky="nsew")
        left = ttk.Frame(paned)
        right = ScrollableFrame(paned)
        paned.add(left, weight=3)
        paned.add(right, weight=2)
        left.columnconfigure(0, weight=1)
        left.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(left, columns=self.columns, show="headings", selectmode="browse")
        setup_tree(
            self.tree,
            {
                "ID": 110,
                "Escala": 90,
                "Entrada/Saída": 70,
                "O.S.": 80,
                "Arquivo": 80,
                "Pax": 55,
                "Extra": 55,
                "Passageiros": 200,
                "Hotel": 220,
                "Idioma": 90,
                "Chegada": 125,
                "Partida": 125,
                "Agência": 120,
            },
        )
        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(left, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        form = ttk.LabelFrame(right.inner, text="Editar grupo", padding=10)
        form.grid(row=0, column=0, sticky="nsew")
        right.inner.columnconfigure(0, weight=1)
        row = 0
        add_labeled_entry(form, row, "ID do grupo", self.vars["group_id"]); row += 1
        add_labeled_entry(form, row, "Escala (DD-MM-YYYY)", self.vars["escala"]); row += 1
        ttk.Checkbutton(form, text="Entrada/Saída", variable=self.vars["in_out"]).grid(row=row, column=0, columnspan=2, sticky="w", pady=3); row += 1
        add_labeled_entry(form, row, "O.S.", self.vars["O.S."]); row += 1
        add_labeled_entry(form, row, "Arquivo", self.vars["file"]); row += 1
        add_labeled_entry(form, row, "Agência", self.vars["agencia"]); row += 1
        add_labeled_entry(form, row, "Pax", self.vars["pax"]); row += 1
        add_labeled_entry(form, row, "Guias externos", self.vars["numero_de_guias_externos"]); row += 1
        ttk.Checkbutton(form, text="pagou pelo guia extra", variable=self.vars["pagou_pelo_guia_extra"]).grid(row=row, column=0, columnspan=2, sticky="w", pady=3); row += 1
        ttk.Label(form, text="Passageiros principais\n(um por linha)").grid(row=row, column=0, sticky="nw", padx=(0, 8), pady=3)
        self.passengers_text = tk.Text(form, height=4, wrap="word")
        self.passengers_text.grid(row=row, column=1, sticky="ew", pady=3); row += 1
        add_labeled_entry(form, row, "Nacionalidade", self.vars["nacionalidade"]); row += 1
        self.hotel_combo = add_labeled_combo(form, row, "Hotel", self.vars["hotel"], []); row += 1
        self.language_combo = add_labeled_combo(form, row, "Idioma", self.vars["idioma"], []); row += 1
        add_labeled_entry(form, row, "Pedágios", self.vars["pedagios"]); row += 1

        sep1 = ttk.Separator(form)
        sep1.grid(row=row, column=0, columnspan=2, sticky="ew", pady=8); row += 1
        ttk.Checkbutton(form, text="Existe voo de chegada", variable=self.vars["voo_chegada_existe"]).grid(row=row, column=0, columnspan=2, sticky="w", pady=3); row += 1
        add_labeled_entry(form, row, "Data de chegada", self.vars["voo_chegada_data"]); row += 1
        add_labeled_entry(form, row, "Número do voo de chegada", self.vars["voo_chegada_numero"]); row += 1
        add_labeled_entry(form, row, "Horário de chegada", self.vars["voo_chegada_horario"]); row += 1

        sep2 = ttk.Separator(form)
        sep2.grid(row=row, column=0, columnspan=2, sticky="ew", pady=8); row += 1
        ttk.Checkbutton(form, text="Existe voo de partida", variable=self.vars["voo_partida_existe"]).grid(row=row, column=0, columnspan=2, sticky="w", pady=3); row += 1
        add_labeled_entry(form, row, "Data de partida", self.vars["voo_partida_data"]); row += 1
        add_labeled_entry(form, row, "Número do voo de partida", self.vars["voo_partida_numero"]); row += 1
        add_labeled_entry(form, row, "Horário de partida", self.vars["voo_partida_horario"]); row += 1

        buttons = ttk.Frame(form)
        buttons.grid(row=row, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        ttk.Button(buttons, text="Novo", command=self.new_item).grid(row=0, column=0, padx=(0, 4))
        ttk.Button(buttons, text="Salvar", command=self.save_item).grid(row=0, column=1, padx=4)
        ttk.Button(buttons, text="Excluir", command=self.delete_item).grid(row=0, column=2, padx=4)

    def refresh_references(self) -> None:
        dates = ["Todas"] + self.app.store.all_escalas()
        self.filter_combo.configure(values=dates)
        self.hotel_combo.configure(values=sorted(self.app.store.hotels))
        self.language_combo.configure(values=self.app.store.all_languages())

    def refresh(self) -> None:
        self.refresh_references()
        self.refresh_tree()

    def refresh_tree(self) -> None:
        clear_tree(self.tree)
        filter_date = normalize_escala(self.filter_var.get())
        for gid, group in sorted(self.app.store.groups.items(), key=lambda item: (date_sort_key(group_escala(item[1])), item[0])):
            ensure_group_defaults(group)
            if filter_date and filter_date != "Todas" and group_escala(group) != filter_date:
                continue
            chegada = format_flight(group, "chegada")
            partida = format_flight(group, "partida")
            values = [
                gid,
                group_escala(group),
                bool_to_text(group.get("in_out")),
                group.get("O.S.", ""),
                group.get("file", ""),
                group.get("pax", ""),
                group.get("numero_de_guias_externos", ""),
                " + ".join(group.get("nomes_dos_passageiros", [])),
                group.get("hotel", ""),
                group.get("idioma", ""),
                chegada,
                partida,
                group.get("agencia", ""),
            ]
            self.tree.insert("", "end", iid=gid, values=values)

    def generate_group_id(self, escala: str) -> str:
        base = f"grupo_{compact_date(escala)}_"
        existing = set(self.app.store.groups)
        n = 1
        while f"{base}{n:03d}" in existing:
            n += 1
        return f"{base}{n:03d}"

    def on_select(self, _event=None) -> None:
        iid = selected_iid(self.tree)
        if not iid:
            return
        self.current_group_id = iid
        group = ensure_group_defaults(self.app.store.groups[iid])
        self.vars["group_id"].set(iid)
        for key, var in self.vars.items():
            if key == "group_id":
                continue
            if isinstance(var, tk.BooleanVar):
                var.set(bool(group.get(key)))
            else:
                var.set("" if group.get(key) is None else str(group.get(key, "")))
        self.passengers_text.delete("1.0", "end")
        self.passengers_text.insert("1.0", "\n".join(group.get("nomes_dos_passageiros", [])))

    def new_item(self) -> None:
        self.current_group_id = None
        default_escala = self.filter_var.get() if self.filter_var.get() and self.filter_var.get() != "Todas" else self.app.current_escala.get()
        default_escala = normalize_escala(default_escala)
        for key, var in self.vars.items():
            if isinstance(var, tk.BooleanVar):
                var.set(False)
            else:
                var.set("")
        self.vars["escala"].set(default_escala)
        self.vars["in_out"].set(True)
        self.vars["pax"].set("1")
        self.vars["numero_de_guias_externos"].set("0")
        self.vars["pagou_pelo_guia_extra"].set(False)
        self.vars["voo_chegada_existe"].set(True)
        self.vars["voo_chegada_data"].set(escala_to_dot(default_escala))
        self.vars["voo_partida_existe"].set(False)
        self.passengers_text.delete("1.0", "end")
        self.tree.selection_remove(self.tree.selection())

    def copy_item(self) -> None:
        iid = selected_iid(self.tree)
        if not iid:
            messagebox.showinfo("Nenhuma seleção", "Selecione um grupo primeiro.")
            return
        self.on_select()
        escala = normalize_escala(self.vars["escala"].get())
        self.current_group_id = None
        self.vars["group_id"].set(self.generate_group_id(escala))
        self.app.set_status("Grupo copiado – revise e salve.")

    def collect_group_from_form(self) -> Tuple[str, Dict[str, Any]]:
        escala = normalize_escala(self.vars["escala"].get())
        if not escala or not is_valid_date_like(escala):
            raise ValueError("A Escala deve estar no formato DD-MM-YYYY.")
        gid = self.vars["group_id"].get().strip() or self.generate_group_id(escala)
        pax = safe_int(self.vars["pax"].get(), -1)
        extra = safe_int(self.vars["numero_de_guias_externos"].get(), -1)
        if pax < 0 or extra < 0:
            raise ValueError("Pax e guias externos devem ser números >= 0.")
        passengers = [line.strip() for line in self.passengers_text.get("1.0", "end").splitlines() if line.strip()]
        chegada_data = self.vars["voo_chegada_data"].get().strip() or escala_to_dot(escala)
        if self.vars["voo_chegada_existe"].get():
            chegada_data = normalize_dot_date(chegada_data)
        partida_data = normalize_dot_date(self.vars["voo_partida_data"].get()) if self.vars["voo_partida_data"].get().strip() else ""
        return gid, {
            "escala": escala,
            "in_out": bool(self.vars["in_out"].get()),
            "pax": pax,
            "pagou_pelo_guia_extra": bool(self.vars["pagou_pelo_guia_extra"].get()),
            "agencia": self.vars["agencia"].get().strip(),
            "O.S.": self.vars["O.S."].get().strip(),
            "numero_de_guias_externos": extra,
            "nomes_dos_passageiros": passengers,
            "nacionalidade": self.vars["nacionalidade"].get().strip(),
            "hotel": self.vars["hotel"].get().strip(),
            "voo_chegada_existe": bool(self.vars["voo_chegada_existe"].get()),
            "voo_chegada_data": chegada_data,
            "voo_chegada_horario": self.vars["voo_chegada_horario"].get().strip(),
            "voo_chegada_numero": self.vars["voo_chegada_numero"].get().strip(),
            "voo_partida_existe": bool(self.vars["voo_partida_existe"].get()),
            "voo_partida_data": partida_data,
            "voo_partida_horario": self.vars["voo_partida_horario"].get().strip(),
            "voo_partida_numero": self.vars["voo_partida_numero"].get().strip(),
            "idioma": self.vars["idioma"].get().strip(),
            "file": self.vars["file"].get().strip(),
            "pedagios": self.vars["pedagios"].get().strip(),
        }

    def save_item(self) -> None:
        try:
            old_id = self.current_group_id
            gid, group = self.collect_group_from_form()
            if old_id and gid != old_id and gid in self.app.store.groups:
                raise ValueError("Este ID do grupo já existe.")
            if not old_id and gid in self.app.store.groups:
                raise ValueError("Este ID do grupo já existe.")
            if old_id and gid != old_id:
                self.app.store.groups.pop(old_id, None)
            self.app.store.groups[gid] = ensure_group_defaults(group)
            self.current_group_id = gid
            self.app.store.save()
            self.app.current_escala.set(group["escala"])
            self.refresh_references()
            if self.filter_var.get() not in ("Todas", group["escala"]):
                self.filter_var.set(group["escala"])
            self.refresh_tree()
            if gid in self.tree.get_children():
                self.tree.selection_set(gid)
                self.tree.see(gid)
            self.app.refresh_references()
            self.app.set_status(f"Grupo salvo: {gid}")
        except Exception as exc:
            show_error("Falha ao salvar grupo", exc)

    def delete_item(self) -> None:
        gid = self.current_group_id or selected_iid(self.tree)
        if not gid:
            return
        if not messagebox.askyesno("Excluir grupo", f"Deseja realmente excluir o grupo {gid}?"):
            return
        self.app.store.groups.pop(gid, None)
        self.current_group_id = None
        self.app.store.save()
        self.refresh()
        self.new_item()
        self.app.refresh_references()
        self.app.set_status(f"Grupo excluído: {gid}")


class SettingsFrame(ttk.Frame):
    def __init__(self, parent, app: TourAssignmentApp):
        super().__init__(parent, padding=16)
        self.app = app
        self.max_one_person_var = tk.StringVar()
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.columnconfigure(1, weight=1)
        ttk.Label(self, text="Arquivo do banco de dados").grid(row=0, column=0, sticky="w", padx=(0, 12), pady=4)
        ttk.Label(self, text=str(self.app.store.path), wraplength=900).grid(row=0, column=1, sticky="w", pady=4)
        add_labeled_entry(self, 1, "Grupo máx. com 1 pessoa", self.max_one_person_var, width=10)
        hint = (
            "Quando pax + guias externos for maior que este valor ou 'pagou pelo guia extra' estiver ativo, "
            "o otimizador planeja Guia e Motorista separadamente."
        )
        ttk.Label(self, text=hint, wraplength=900).grid(row=2, column=1, sticky="w", pady=(0, 10))
        ttk.Button(self, text="Salvar configurações", command=self.save_settings).grid(row=3, column=1, sticky="w", pady=8)
        deps = (
            "Pacotes recomendados:\n"
            "  pip install scipy numpy openpyxl\n\n"
            "scipy/numpy = otimização exata. Se esses pacotes não estiverem instalados, o app usa um fallback simples.\n"
            "openpyxl = exportação em Excel. CSV e JSON funcionam sem pacotes adicionais."
        )
        ttk.Label(self, text=deps, wraplength=900).grid(row=4, column=0, columnspan=2, sticky="w", pady=(20, 0))

    def refresh(self) -> None:
        self.max_one_person_var.set(str(self.app.store.settings.get("maximum_group_size_with_only_one_person", 7)))

    def save_settings(self) -> None:
        value = safe_int(self.max_one_person_var.get(), -1)
        if value < 0:
            messagebox.showwarning("Verificar valor", "Informe um número >= 0.")
            return
        self.app.store.settings["maximum_group_size_with_only_one_person"] = value
        self.app.store.save()
        self.app.set_status("Configurações salvas.")


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------


def export_csv(rows: List[List[Any]], target: Path) -> None:
    with target.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerows(rows)


def export_json(rows: List[List[Any]], target: Path) -> None:
    with target.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def export_xlsx(rows: List[List[Any]], target: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # pragma: no cover - abhängig vom Zielsystem
        raise RuntimeError("Para exportar para Excel, instale: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Escala"
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, _header in enumerate(rows[0], start=1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(target)


# ---------------------------------------------------------------------------
# Start
# ---------------------------------------------------------------------------


def main() -> None:
    app = TourAssignmentApp()
    app.mainloop()


if __name__ == "__main__":
    main()

